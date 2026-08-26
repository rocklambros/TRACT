"""Rebuild the tracked fingerprint set for restricted framework sources.

    python -m scripts.build_licensed_fingerprints
    python -m scripts.build_licensed_fingerprints --check

Writes `tests/fixtures/licensed_text_fingerprints.json`, the committed input to
`tests/test_licensed_text_not_tracked.py`. That gate used to read the licensed
source out of gitignored `data/raw/`, so on a fresh clone and in CI it skipped
and the skip reported green.

Scope is `tract.licensing.fingerprinted_framework_ids()`: every framework that
routes to the gitignored overlay, less a named and justified exclusion. It was
RESTRICTED_FRAMEWORK_IDS, which covered ISO/IEC 27001 and ETSI and left CSA CCM
and DSOMM contributing nothing, so a quotation from either would have walked
past the gate that exists to catch exactly that.

RE-PIN THE FINGERPRINTS WHENEVER A SOURCE CHANGES. Every entry is derived from
one document's bytes, and the sha256 of that document is recorded alongside the
entry. Re-fetching a source, correcting a parse, or moving a framework into the
overlay tier all require running this script on a checkout that holds the
sources under data/raw/, then committing the regenerated file. `--check`
reports drift without writing, so a reviewer can tell a stale fingerprint file
from a current one.

The output carries hashes only. There is no free-text field in the schema, and
tests/test_licensed_text_not_tracked.py asserts that, so this script cannot
leak the text it reads.

Owner: TRACT.
"""
from __future__ import annotations

import argparse
import hashlib
import logging
import re
import sys
from collections.abc import Callable
from pathlib import Path
from typing import Any, Final

from tract.config import OVERLAY_FRAMEWORK_IDS, RAW_FRAMEWORKS_DIR
from tract.io import atomic_write_json
from tract.licensing import (
    FINGERPRINT_ALGORITHM,
    FINGERPRINT_EXCLUDED_FRAMEWORK_IDS,
    FINGERPRINT_GENERATOR,
    FINGERPRINT_HEX_CHARS,
    FINGERPRINT_PATH,
    FINGERPRINT_SALT,
    NGRAM_WORDS,
    LicensedFingerprints,
    fingerprint_ngrams,
    fingerprinted_framework_ids,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

CHUNK_BYTES: Final[int] = 1 << 16

# A row of ISO's Annex A table: "| 5.1 | <title> | Control <statement> |".
_ANNEX_A_CLAUSE: Final[re.Pattern[str]] = re.compile(r"^\d+\.\d+$")
_ANNEX_A_CELLS: Final[int] = 3


def _iso_27001_statements(path: Path) -> list[str]:
    """ISO/IEC 27001:2022 Annex A control statements, title column excluded.

    The titles are deliberately left out. Section names like "Privacy and
    protection of personal identifiable information (PII)" reach this project
    through OpenCRE's public link dump and are already tracked under
    data/training/hub_links*. They are not the normative text at issue. What
    must never enter git is the requirement statement, which the source table
    marks with a leading "Control".
    """
    statements: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.startswith("|"):
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if len(cells) != _ANNEX_A_CELLS or not _ANNEX_A_CLAUSE.match(cells[0]):
            continue
        statements.append(cells[2])
    if not statements:
        raise ValueError(
            f"{path}: no Annex A table rows matched. The extraction changed "
            f"shape and the fingerprints would be silently empty."
        )
    return statements


# The page header ETSI repeats on every page: "13 ETSI GR SAI 005 V1.1.1 ...".
# It is furniture, and leaving it in interleaves five stray words into the word
# stream at every page break, which breaks windows that span one.
_ETSI_PAGE_HEADER: Final[re.Pattern[str]] = re.compile(r"^\d+\s+ETSI\s+GR\s+")
# Table-of-contents rows, recognised by their dot leaders.
_TOC_LEADER: Final[re.Pattern[str]] = re.compile(r"\.{5,}")
# The clause that opens the bibliography and the clause that closes it.
_ETSI_REFERENCES_START: Final[re.Pattern[str]] = re.compile(r"^2\s+References\s*$")
_ETSI_REFERENCES_END: Final[re.Pattern[str]] = re.compile(r"^3\s+Definition\b")


def _etsi_normative_text(path: Path) -> list[str]:
    """ETSI GR SAI 005's own prose, bibliography and contents page excluded.

    ETSI has no parser yet, so there is no statement column to isolate the way
    ISO's has. The whole deliverable is fingerprinted instead, front matter and
    copyright notification included, so this repository cannot quote that
    notice at length either.

    Two parts are deliberately left out because they are not ETSI's authored
    content, and including them produced false positives that were measured
    rather than imagined:

      - Clause 2, the informative references. It is a list of other people's
        paper titles and author names. NIST AI 100-2 and the OWASP AI Exchange
        cite several of the same papers, so their tracked JSON reproduced
        author lists such as the one behind [i.2] word for word. Flagging that
        as an ETSI quotation is wrong on the facts.
      - The contents page, whose dot leaders make heading text run together
        into windows that exist nowhere in the body.

    Raises:
        ValueError: neither clause boundary was found, which means the document
            structure changed and the bibliography would silently be back in.
    """
    import pdfplumber  # imported here so the gate never needs it, only the build

    with pdfplumber.open(path) as pdf:
        pages = [page.extract_text() or "" for page in pdf.pages]
    raw = "\n".join(pages)
    if not raw.strip():
        raise ValueError(
            f"{path}: pdfplumber extracted no text. A scanned or protected PDF "
            f"would produce an empty fingerprint set that looks like success."
        )

    kept: list[str] = []
    in_references = False
    saw_start = saw_end = False
    for line in raw.splitlines():
        stripped = line.strip()
        if _ETSI_REFERENCES_START.match(stripped):
            in_references, saw_start = True, True
            continue
        if in_references and _ETSI_REFERENCES_END.match(stripped):
            in_references, saw_end = False, True
        if in_references:
            continue
        if _ETSI_PAGE_HEADER.match(stripped) or _TOC_LEADER.search(stripped):
            continue
        kept.append(stripped)

    if not (saw_start and saw_end):
        raise ValueError(
            f"{path}: could not bound the references clause "
            f"(start={saw_start}, end={saw_end}). The document structure "
            f"changed. Fix the boundaries rather than fingerprinting the "
            f"bibliography, which produces false positives on shared citations."
        )

    # One unit, not one per page. A sentence that runs across a page break is
    # still one sentence, and a quotation of it would carry the break too.
    return ["\n".join(kept)]


# Cell values arrive with the workbook's own line breaks and runs of spaces.
# The parser collapses both before it builds a Control, so collapsing them here
# too keeps the fingerprinted wording identical to the wording that could leak.
_WHITESPACE: Final[re.Pattern[str]] = re.compile(r"\s+")


def _check_pinned_name(path: Path, expected: str, owner: str) -> None:
    """Refuse a source file whose parser has since been re-pinned elsewhere.

    The filenames in _EXTRACTORS and the ones the parsers pin are two records
    of one fact. A re-pin that moves only the parser leaves this script reading
    a superseded document, and the fingerprints would then describe text that
    no longer reaches the corpus.

    Raises:
        ValueError: the staged filename is not the one *owner* pins.
    """
    if path.name == expected:
        return
    raise ValueError(
        f"{path.name} is not {expected!r}, which {owner} pins. Fingerprints "
        f"built from a superseded document describe text that no longer "
        f"reaches the corpus. Update _EXTRACTORS and re-measure together."
    )


def _csa_ccm_specifications(path: Path) -> list[str]:
    """CSA CCM v4.1.0 control specifications, every other column excluded.

    Read from the same workbook, the same sheet and the same column
    parsers/parse_csa_ccm.py reads, with the constants imported from it rather
    than restated, so a re-pin of the workbook moves both together.

    The Control Title and Control Domain columns are left out for the reason
    ISO's titles are: OpenCRE's public link dump carries all 29 CCM section
    names, they are already tracked in data/training/hub_links*, and the merge
    keeps overlay frameworks' titles in the tracked corpus on purpose. The
    normative text at issue is the specification.

    Domain rows contribute nothing here. A CCM domain has no text of its own,
    and the statement the corpus carries for one is a list of member titles
    that parse_csa_ccm.py assembled and marked synthetic. Fingerprinting
    TRACT's own assembled text would flag TRACT's own output as a CSA
    quotation.

    Raises:
        ValueError: the header row is absent or reordered, or no specification
            rows matched. Column order is the whole assumption: swapped columns
            still yield 207 rows, with titles where the specifications should
            be, and every count check downstream still passes.
    """
    import openpyxl  # imported here so the gate never needs it, only the build

    from parsers.parse_csa_ccm import EXPECTED_HEADER, SHEET_NAME, WORKBOOK_NAME

    _check_pinned_name(path, WORKBOOK_NAME, "parsers/parse_csa_ccm.py")
    specification_column = EXPECTED_HEADER.index("Control Specification")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"{path}: no {SHEET_NAME!r} sheet, only "
                f"{workbook.sheetnames}. The CAIQ sheet is the "
                f"self-assessment questionnaire and is not the controls."
            )
        rows = [
            tuple(
                "" if cell is None else _WHITESPACE.sub(" ", str(cell)).strip()
                for cell in (list(row) + [None] * 4)[:4]
            )
            for row in workbook[SHEET_NAME].iter_rows(values_only=True)
        ]
    finally:
        workbook.close()

    headers = sum(1 for row in rows if row == EXPECTED_HEADER)
    if headers != 1:
        raise ValueError(
            f"{path}: the {SHEET_NAME} sheet carries {headers} header rows "
            f"equal to {list(EXPECTED_HEADER)}, expected 1. Without that "
            f"anchor there is nothing to say column "
            f"{specification_column} still holds specifications, and "
            f"fingerprinting the title column produces false positives on "
            f"tracked files that carry nothing licensed."
        )

    statements = [
        row[specification_column]
        for row in rows
        if row != EXPECTED_HEADER and row[2] and row[specification_column]
    ]
    if not statements:
        raise ValueError(
            f"{path}: no control specification rows matched. The sheet changed "
            f"shape and the fingerprints would be silently empty."
        )
    return statements


def _dsomm_statements(path: Path) -> list[str]:
    """OWASP DSOMM activity statements, activity and dimension names excluded.

    Read from the archive member and the field set parsers/parse_dsomm.py
    reads, joined in the same order, so what is fingerprinted is the text that
    would leak if the parsed corpus reached git.

    Joined rather than fingerprinted per field, unlike the per-statement units
    elsewhere in this file. The join is what the parser writes into one
    `description`, so a window spanning the seam between `risk` and `measure`
    is text somebody could quote from the corpus, not an artifact of
    concatenating two unrelated records.

    The activity name is the YAML key and is the control's title. It is left
    out for the same reason ISO's and CCM's titles are: `dimension` and
    `sub_dimension` are what OpenCRE's public link dump puts in `section_name`
    for all 214 DSOMM links, and they are already tracked.

    Raises:
        ValueError: the model member is absent or ambiguous, the stream is not
            a meta document followed by a model mapping, or no activity
            produced text.
    """
    import zipfile
    from io import BytesIO

    import yaml

    from parsers.parse_dsomm import ARCHIVE_NAME, MODEL_SUFFIX, STATEMENT_FIELDS

    _check_pinned_name(path, ARCHIVE_NAME, "parsers/parse_dsomm.py")
    with zipfile.ZipFile(BytesIO(path.read_bytes())) as archive:
        names = [n for n in archive.namelist() if n.endswith(MODEL_SUFFIX)]
        if len(names) != 1:
            raise ValueError(
                f"{path}: expected exactly one {MODEL_SUFFIX} member, found "
                f"{names}. The generated file is what flattens the "
                f"per-subdimension YAMLs into one document."
            )
        raw = archive.read(names[0]).decode("utf-8")

    documents = list(yaml.safe_load_all(raw))
    if len(documents) != 2 or not isinstance(documents[1], dict):
        raise ValueError(
            f"{path}: {MODEL_SUFFIX} is not a meta document followed by a "
            f"model mapping (got {len(documents)} document(s)). The layout "
            f"changed and the fingerprints would be silently empty."
        )

    statements: list[str] = []
    for sub_dimensions in documents[1].values():
        for activities in sub_dimensions.values():
            for body in activities.values():
                parts = [
                    str(body.get(field) or "").strip()
                    for field in STATEMENT_FIELDS
                ]
                statement = "\n\n".join(part for part in parts if part)
                if statement:
                    statements.append(statement)
    if not statements:
        raise ValueError(
            f"{path}: no activity carried text in any of {STATEMENT_FIELDS}. "
            f"The schema changed and the fingerprints would be silently empty."
        )
    return statements


# One extractor per fingerprinted framework. Keyed on framework_id so adding a
# framework to the overlay tier without deciding what its statement text is
# raises here instead of producing a fingerprint file with a hole.
_EXTRACTORS: Final[dict[str, tuple[str, Callable[[Path], list[str]]]]] = {
    "csa_ccm": ("CCMv4.1.0-generated_at_2026_01_13.xlsx", _csa_ccm_specifications),
    "dsomm": ("dsomm_data.zip", _dsomm_statements),
    "iso_27001": ("ISO_IEC_27001_2022_en.md", _iso_27001_statements),
    "etsi": ("etsi_gr_sai005_v010101p.pdf", _etsi_normative_text),
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(CHUNK_BYTES), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build() -> dict[str, Any]:
    """Read every fingerprinted source and return the fingerprint document.

    Raises:
        KeyError: a framework is in scope but has no extractor here.
        FileNotFoundError: a source in scope is not staged under data/raw/.
        ValueError: an excluded framework id names no known framework.
    """
    in_scope = fingerprinted_framework_ids()
    # Every OVERLAY framework needs an extractor, not merely every in-scope one.
    # A deferred framework whose extractor was never written or was deleted
    # would make the deferral irreversible in practice, and the reason a
    # framework is deferred is a licence question that can be answered, not a
    # missing-code question. csa_ccm's extractor is registered and measured for
    # exactly this reason.
    missing_extractors = sorted(set(OVERLAY_FRAMEWORK_IDS) - set(_EXTRACTORS))
    if missing_extractors:
        raise KeyError(
            f"No fingerprint extractor for overlay framework(s) "
            f"{missing_extractors}. Add one to _EXTRACTORS naming the staged "
            f"document and the function that isolates its statement text. "
            f"Naming it in FINGERPRINT_EXCLUDED_FRAMEWORK_IDS defers the gate, "
            f"which still needs the extractor to exist so the deferral can be "
            f"reversed by a decision rather than by new code."
        )

    documents: list[dict[str, Any]] = []
    fingerprints: set[str] = set()
    for framework_id in sorted(in_scope):
        filename, extract = _EXTRACTORS[framework_id]
        path = RAW_FRAMEWORKS_DIR / framework_id / filename
        if not path.exists():
            raise FileNotFoundError(
                f"{path} is not staged. This script has to run on a checkout "
                f"that holds every restricted source under data/raw/."
            )
        # Fingerprinted per unit, never across the join between two of them.
        # A window spanning the end of one ISO statement and the start of the
        # next is not text anyone could quote, and storing it only widens the
        # surface for an accidental match.
        units = extract(path)
        grams = [gram for unit in units for gram in fingerprint_ngrams(unit)]
        fingerprints.update(grams)
        documents.append({
            "framework_id": framework_id,
            "filename": filename,
            "source_sha256": _sha256(path),
            "ngram_count": len(grams),
        })
        logger.info(
            "%s/%s: %d unit(s), %d words, %d n-grams", framework_id, filename,
            len(units), sum(len(unit.split()) for unit in units), len(grams),
        )

    return {
        "salt": FINGERPRINT_SALT,
        "algorithm": FINGERPRINT_ALGORITHM,
        "ngram_words": NGRAM_WORDS,
        "hash_hex_chars": FINGERPRINT_HEX_CHARS,
        "generator": FINGERPRINT_GENERATOR,
        # Ids only, never a reason. See FINGERPRINT_TOP_LEVEL_KEYS.
        "deferred_framework_ids": sorted(FINGERPRINT_EXCLUDED_FRAMEWORK_IDS),
        "documents": documents,
        # Sorted so a rebuild from unchanged sources is byte-identical and a
        # diff on this file means the sources actually moved.
        "fingerprints": sorted(fingerprints),
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Rebuild tests/fixtures/licensed_text_fingerprints.json",
    )
    parser.add_argument(
        "--check", action="store_true",
        help="report drift against the committed file without writing it",
    )
    args = parser.parse_args()

    document = build()
    logger.info(
        "Built %d fingerprints from %d document(s) at n=%d words",
        len(document["fingerprints"]), len(document["documents"]), NGRAM_WORDS,
    )

    if args.check:
        try:
            committed = LicensedFingerprints.load(FINGERPRINT_PATH)
        except (FileNotFoundError, ValueError) as exc:
            logger.error("committed fingerprint file unusable: %s", exc)
            return 1
        built = frozenset(document["fingerprints"])
        if committed.fingerprints == built:
            logger.info("%s is current", FINGERPRINT_PATH)
            return 0
        logger.error(
            "%s is STALE: %d fingerprints only in the committed file, %d only "
            "in the rebuild. Re-run without --check and commit the result.",
            FINGERPRINT_PATH,
            len(committed.fingerprints - built), len(built - committed.fingerprints),
        )
        return 1

    atomic_write_json(document, FINGERPRINT_PATH)
    logger.info("Wrote %s", FINGERPRINT_PATH)
    return 0


if __name__ == "__main__":
    sys.exit(main())
