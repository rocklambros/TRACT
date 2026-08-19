"""Parser for the CSA Cloud Controls Matrix v4.1.0.

Not the AI Controls Matrix. This workbook's title cell reads CLOUD CONTROLS
MATRIX v4.1.0 and it is a different framework from csa_aicm, which has 243
controls and no CRE links.

The CCM sheet is a flat four-column table with two row types interleaved. A
control row populates all four columns; a domain header populates only column A
in the form "<Full Name> - <CODE>". Three column-A-only rows are neither: the
workbook title, the End of Standard trailer, and the copyright paragraph. And
one all-four-columns row is the header itself, which is why a naive count gives
208 control rows where there are 207. [measured]

Both granularities are emitted, because OpenCRE links both: 14 of the 29
curated links target a control id and 15 target a bare domain code. [measured]

A domain's statement is the ordered list of its member control TITLES, not the
concatenation of their specifications. Measured on this workbook, concatenating
specifications runs 1,022 to 4,292 characters, which puts 8 of 17 domain
anchors over MAX_ANCHOR_CHARS and makes all 17 open with their own first member
control. That would put 17 near-duplicate pairs into a 29-link framework. The
title list runs 163 to 596 characters, exceeds nothing, and opens nothing.

That list is assembled text, not a rule anyone wrote, and 14 of the 29 curated
links land on one. So every aggregate carries text_origin = synthetic and every
aggregate gets a repair-audit record carrying its before and after as text.
honest_prose_fraction cannot tell the difference and would count a list of
subjects as a control statement.

The seven curated links that still use v4.0's IVS-* ids need no rename map:
their section_name matches the corresponding I&S-* control's title exactly, and
ProseIndex resolves title before id. [measured] Three AIS rows carry v4.0
titles for controls v4.1.0 renamed, and those are declared as alternates. See
OPENCRE_TITLE_VARIANTS for why declaring them is not cosmetic.

One divergence is real and is recorded rather than repaired. See
KNOWN_DIVERGENCES.
"""
from __future__ import annotations

import hashlib
import logging
import re
from collections import defaultdict
from collections.abc import Mapping
from io import BytesIO
from typing import ClassVar, Final, NamedTuple

import openpyxl

from tract.config import DESCRIPTION_MAX_LENGTH
from tract.corpus_report import SYNTHETIC_TEXT_ORIGIN, TEXT_ORIGIN_METADATA_KEY
from tract.parsers.base import BaseParser
from tract.schema import Control

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

WORKBOOK_NAME: Final[str] = "CCMv4.1.0-generated_at_2026_01_13.xlsx"
SOURCE_SHA256: Final[str] = (
    "5e721628c8ab297bdbd355afa4c01699971fcbb9cb16802ccb9d42c7176ab32b"
)
SHEET_NAME: Final[str] = "CCM"

# "Cryptography, Encryption & Key Management - CEK". The code is 2 to 5
# characters of uppercase letters and ampersands, which covers A&A and I&S.
DOMAIN_HEADER: Final[re.Pattern[str]] = re.compile(r"^(.+?)\s+-\s+([A-Z&]{2,5})$")

# Exactly one row must equal this, and column order is the single assumption
# the whole parse rests on. Swapped columns still yield 207 control rows and 17
# domains with title and specification exchanged, so every count check passes
# and nothing else in this file can see it.
EXPECTED_HEADER: Final[tuple[str, str, str, str]] = (
    "Control Domain", "Control Title", "Control ID", "Control Specification",
)

# Column-A-only rows that are not domain headers. Anything else column-A-only
# raises, because a skipped row is a domain that quietly stops existing.
TRAILERS: Final[tuple[str, ...]] = ("End of Standard", "© Copyright",
                                    "(c) Copyright")

# Controls v4.1.0 renamed, keyed to the v4.0 title OpenCRE's curated links
# still spell. Declared as alt_titles so all three links resolve through the
# title channel the curator wrote.
#
# Not cosmetic. Without this table AIS-04's link answers by id, and the id-side
# wrong-anchor detector then compares OpenCRE's "Secure Application Design and
# Development" against v4.1.0's "Secure Application Development Lifecycle",
# finds neither string inside the other, and flags it. Measured: the framework
# reports wrong_anchor_risk 2 without this table and 1 with it, against the 1
# pre-registered in JOIN_WRONG_ANCHOR_BUDGET. The anchor was provably right
# either way -- the section_id IS AIS-04 -- so the flag was a fact about a
# stale spelling rather than about the anchor.
#
# AIS-05 and AIS-06 are the same rename and are declared for the same reason
# even though the detector's substring test happens to let them through:
# v4.1.0's title is a substring of OpenCRE's in those two cases and not in
# AIS-04's. Declaring the set the source renamed, rather than the subset one
# detector flags, is what keeps this a derived table instead of an allowlist.
#
# tests/test_parse_csa_ccm.py derives this same set from the tracked link file
# and from the parsed workbook, so a stale entry and a new mismatch both fail.
OPENCRE_TITLE_VARIANTS: Final[Mapping[str, tuple[str, ...]]] = {
    "AIS-04": ("Secure Application Design and Development",),
    "AIS-05": ("Automated Application Security Testing",),
    "AIS-06": ("Automated Secure Application Deployment",),
}


class Divergence(NamedTuple):
    """A curated link whose two sides name different controls, on purpose."""

    opencre_section_name: str
    resolved_to: str
    reason: str


# The one curated row whose section_name names a control while its section_id
# names that control's domain. Hand-verified against the pinned workbook and
# against the link's CRE target. Recorded, never repaired.
KNOWN_DIVERGENCES: Final[Mapping[str, Divergence]] = {
    "IPY": Divergence(
        opencre_section_name=(
            "Interoperability and portability policy and procedures"
        ),
        resolved_to="IPY-01",
        reason=(
            "OpenCRE's section_id is the IPY domain and its section_name is "
            "control IPY-01's title, which is also the name of its CRE target "
            "847-247. [measured] The title channel therefore answers with "
            "IPY-01. That is left standing: IPY-01's 462-character "
            "specification is a rule about the subject the CRE names, and the "
            "IPY domain's statement is a four-item list of subjects this "
            "parser assembled. The corpus report reads wrong_anchor_risk = 1 "
            "for this framework on purpose."
        ),
    ),
}

_WHITESPACE: Final[re.Pattern[str]] = re.compile(r"\s+")

# Room for the sentence-ending period a domain statement adds. The guard below
# refuses at the limit rather than near it, so this is the margin the message
# quotes, not a second threshold.
_DESCRIPTION_LIMIT: Final[int] = DESCRIPTION_MAX_LENGTH


class CsaCcmParser(BaseParser):
    framework_id: ClassVar[str] = "csa_ccm"
    # Matches the curated links' standard_name exactly; no alias entry exists
    # or is needed. [measured]
    framework_name: ClassVar[str] = "Cloud Controls Matrix"
    version: ClassVar[str] = "4.1.0"
    source_url: ClassVar[str] = (
        "https://cloudsecurityalliance.org/artifacts/cloud-controls-matrix-v4/"
    )
    # The dominant unit. 17 domain aggregates ride alongside the 207 controls
    # because 15 of the 29 curated links target a bare domain code, and a
    # framework that emitted only controls would leave those 15 unresolvable.
    mapping_unit_level: ClassVar[str] = "control"
    # 207 control rows plus 17 domains. The 208th all-four-columns row is the
    # sheet header. [measured]
    expected_count: ClassVar[int] = 224
    # COUNT_TOLERANCE is 10%, so the band around 224 is 202 to 246 and a parse
    # that lost 22 controls would write in silence. These two are the
    # structural check that beats the band. Overridable so a synthetic
    # workbook can drive parse() where data/raw is absent.
    expected_control_rows: ClassVar[int] = 207
    expected_domains: ClassVar[int] = 17
    fetched_date: ClassVar[str] = "2026-08-15"
    # 222 of 224 units clear HONEST_PROSE_MIN_CHARS. IAM-07's specification is
    # 58 characters and STA-06's is 43, giving 0.9911. A floor of 1.0 refuses
    # to write on correct output; 0.99 passes at 222/224 and fails at 221/224
    # (0.9866), so it can fail in both directions and its margin is one
    # control. [measured]
    min_prose_fraction: ClassVar[float] = 0.99
    expected_sha256: ClassVar[str | None] = SOURCE_SHA256
    # Class-level, and read through `self` in parse(), so a fixture workbook
    # that carries three domains can narrow a table naming controls it does
    # not have. A classmethod cannot see an instance attribute, which is why
    # both are passed down explicitly rather than read off `cls`.
    title_variants: ClassVar[Mapping[str, tuple[str, ...]]] = (
        OPENCRE_TITLE_VARIANTS
    )
    known_divergences: ClassVar[Mapping[str, Divergence]] = KNOWN_DIVERGENCES

    def parse(self) -> list[Control]:
        payload = self.read_source_bytes(WORKBOOK_NAME)
        self._check_digest(payload)
        rows = self._read_sheet(payload)
        controls = self.rows_to_controls(rows, self.title_variants)
        self._check_shape(controls)
        self.write_repair_audit(
            self.domain_audit_records(controls, self.known_divergences)
        )
        domains = sum(1 for c in controls if c.hierarchy_level == "domain")
        logger.info(
            "%s: %d controls and %d domains, %d synthesised domain statement(s)",
            self.framework_id, len(controls) - domains, domains, domains,
        )
        return controls

    def _check_digest(self, payload: bytes) -> None:
        """Refuse a workbook that is not the pinned one.

        Raises:
            ValueError: If the digest does not match `expected_sha256`.
        """
        if self.expected_sha256 is None:
            return
        actual = hashlib.sha256(payload).hexdigest()
        if actual == self.expected_sha256:
            return
        raise ValueError(
            f"{self.framework_id}: {WORKBOOK_NAME} has sha256 {actual}, not "
            f"the pinned {self.expected_sha256}. expected_count of 224, the "
            f"domain-statement length measurements that chose member titles "
            f"over member specifications, the prose floor and the join ceiling "
            f"were all taken against the pinned bytes. The workbook is "
            f"registration-walled and hand-staged, so it cannot be re-fetched: "
            f"re-measure before moving the pin."
        )

    def _check_shape(self, controls: list[Control]) -> None:
        """Refuse a parse whose row shape moved, before the band can hide it.

        Raises:
            ValueError: If the control-row or domain count differs from the
                declared shape.
        """
        found_controls = sum(1 for c in controls if c.hierarchy_level == "control")
        found_domains = sum(1 for c in controls if c.hierarchy_level == "domain")
        if found_controls != self.expected_control_rows:
            raise ValueError(
                f"{self.framework_id}: {found_controls} control rows, "
                f"expected {self.expected_control_rows}. COUNT_TOLERANCE puts "
                f"the band around the 224 total at 202 to 246, so a loss of up "
                f"to 22 controls would write without a word."
            )
        if found_domains != self.expected_domains:
            raise ValueError(
                f"{self.framework_id}: {found_domains} domains, expected "
                f"{self.expected_domains}. Fifteen of the 29 curated links "
                f"target a bare domain code, so a lost domain is a lost link."
            )

    def _read_sheet(self, payload: bytes) -> list[tuple[str, str, str, str]]:
        """The CCM sheet's first four columns, as stripped strings.

        Read from the same bytes BaseParser hashed, not from a second open of
        the path. openpyxl takes a file object, so there is no reason to leave
        a window in which the digest gate and the parse could see different
        bytes. openpyxl.DEFUSEDXML reads True in this environment, so the
        workbook's XML is parsed with entity handling hardened. [measured]

        Raises:
            ValueError: If the CCM sheet is absent.
        """
        workbook = openpyxl.load_workbook(
            BytesIO(payload), read_only=True, data_only=True,
        )
        try:
            if SHEET_NAME not in workbook.sheetnames:
                raise ValueError(
                    f"{self.framework_id}: {WORKBOOK_NAME} has no "
                    f"{SHEET_NAME!r} sheet, only {workbook.sheetnames}. The "
                    f"CAIQ sheet is the self-assessment questionnaire and is "
                    f"not the controls."
                )
            rows: list[tuple[str, str, str, str]] = []
            for row in workbook[SHEET_NAME].iter_rows(values_only=True):
                # None-checked rather than falsy-checked: `str(cell or "")`
                # turns a numeric zero into an empty cell, and a cell that
                # empties itself changes which row type a row is.
                cells = [
                    "" if cell is None else _WHITESPACE.sub(" ", str(cell)).strip()
                    for cell in (list(row) + [None] * 4)[:4]
                ]
                rows.append((cells[0], cells[1], cells[2], cells[3]))
            return rows
        finally:
            workbook.close()

    @classmethod
    def rows_to_controls(
        cls,
        rows: list[tuple[str, str, str, str]],
        title_variants: Mapping[str, tuple[str, ...]] | None = None,
    ) -> list[Control]:
        """Controls then domains, from the sheet's interleaved row types.

        `title_variants` defaults to the class table. It is a parameter rather
        than a `cls.title_variants` read because a fixture workbook carries a
        few domains and cannot satisfy a table naming controls it does not
        have, and an instance attribute is invisible to a classmethod.

        Raises:
            ValueError: If the header row is absent or reordered, a column-A
                row is neither a domain header nor a known trailer, two rows
                claim one control id, a domain header has no control rows
                under it, or any statement would reach DESCRIPTION_MAX_LENGTH.
        """
        variants = cls.title_variants if title_variants is None else title_variants
        cls._check_header(rows)

        controls: list[Control] = []
        domains: list[tuple[str, str]] = []
        members: dict[str, list[str]] = defaultdict(list)
        seen: set[str] = set()
        current_code = ""
        current_name = ""

        for first, title, control_id, specification in rows:
            if control_id and specification:
                if (first, title, control_id, specification) == EXPECTED_HEADER:
                    continue
                if control_id in seen:
                    raise ValueError(
                        f"csa_ccm: a second row claims control id "
                        f"{control_id!r}, titled {title!r}. Both would be "
                        f"indexed and the later would win, so one control "
                        f"would become unreachable through the id channel "
                        f"with nothing to say which."
                    )
                seen.add(control_id)
                controls.append(Control(
                    control_id=control_id,
                    title=title,
                    description=specification,
                    hierarchy_level="control",
                    parent_id=current_code or None,
                    parent_name=current_name or first or None,
                ))
                members[current_code].append(title)
                continue

            if not first or title or control_id or specification:
                continue
            if first.startswith(TRAILERS):
                continue
            header = DOMAIN_HEADER.match(first)
            if header is None:
                raise ValueError(
                    f"csa_ccm: column-A-only row {first[:120]!r} is neither a "
                    f"domain header of the form '<Name> - <CODE>' nor one of "
                    f"the known trailers {list(TRAILERS)}. Skipping it would "
                    f"lose a domain, and 15 of the 29 curated links target a "
                    f"bare domain code."
                )
            current_name, current_code = header.group(1), header.group(2)
            domains.append((current_code, current_name))

        controls = cls._apply_title_variants(controls, variants)
        controls += cls._domain_controls(domains, members)
        cls._check_description_budget(controls)
        return controls

    @staticmethod
    def _check_header(rows: list[tuple[str, str, str, str]]) -> None:
        """Refuse a sheet whose column order moved.

        Raises:
            ValueError: If the sheet does not carry exactly one header row.
        """
        found = sum(1 for row in rows if row == EXPECTED_HEADER)
        if found == 1:
            return
        raise ValueError(
            f"csa_ccm: the CCM sheet carries {found} header rows equal to "
            f"{list(EXPECTED_HEADER)}, expected 1. Column order is the one "
            f"assumption this parse rests on: swapped columns still yield the "
            f"declared control and domain counts, with title and "
            f"specification exchanged, so nothing else here can see it."
        )

    @staticmethod
    def _apply_title_variants(
        controls: list[Control],
        variants: Mapping[str, tuple[str, ...]],
    ) -> list[Control]:
        """Attach the v4.0 spellings OpenCRE's curated links still use.

        Raises:
            ValueError: If a declared control is absent, or a declared variant
                is empty or restates the control's own title.
        """
        by_id = {control.control_id: control for control in controls}
        for control_id in sorted(variants):
            control = by_id.get(control_id)
            if control is None:
                raise ValueError(
                    f"csa_ccm: the title variant table names no control "
                    f"{control_id!r}. A variant for a control that does not "
                    f"exist reaches nothing and still reads as a live "
                    f"alternate."
                )
            for variant in variants[control_id]:
                if not variant.strip():
                    raise ValueError(
                        f"csa_ccm: control {control_id} declares an empty "
                        f"title variant. An empty key can never be looked up."
                    )
                if variant.strip().lower() == control.title.strip().lower():
                    raise ValueError(
                        f"csa_ccm: control {control_id} declares the title "
                        f"variant {variant!r}, which is already the control's "
                        f"own title. ProseIndex indexes real titles first and "
                        f"never lets an alternate displace one, so this entry "
                        f"is dead."
                    )
            metadata: dict[str, str | list[str]] = dict(control.metadata or {})
            metadata["alt_titles"] = list(variants[control_id])
            by_id[control_id] = control.model_copy(update={"metadata": metadata})
        return [by_id[control.control_id] for control in controls]

    @staticmethod
    def _domain_controls(
        domains: list[tuple[str, str]], members: Mapping[str, list[str]],
    ) -> list[Control]:
        """One mapping unit per domain, stated as its members' subjects.

        Raises:
            ValueError: If a domain owns no controls, which means the row
                ordering changed and every domain statement built from that
                ordering is attached to the wrong domain.
        """
        built: list[Control] = []
        for code, name in domains:
            titles = members.get(code, [])
            if not titles:
                raise ValueError(
                    f"csa_ccm: domain {code} has no controls under it. Domain "
                    f"membership comes from row order, so an empty domain "
                    f"means the sheet was reordered and the domains that do "
                    f"have members may have the wrong ones."
                )
            built.append(Control(
                control_id=code,
                title=name,
                description=". ".join(titles) + ".",
                hierarchy_level="domain",
                metadata={
                    # Titles, not ids. An earlier draft named this key
                    # member_ids and held titles, which is the kind of thing a
                    # reader trusts once and never rechecks.
                    "member_titles": list(titles),
                    TEXT_ORIGIN_METADATA_KEY: SYNTHETIC_TEXT_ORIGIN,
                },
            ))
        return built

    @staticmethod
    def _check_description_budget(controls: list[Control]) -> None:
        """Refuse a statement BaseParser._sanitize_control would rewrite.

        Ruling R14: that function truncates `description` past
        DESCRIPTION_MAX_LENGTH and writes the untruncated text into
        `full_text`, discarding whatever the parser put there. It does not
        fire on the pinned workbook, whose longest control specification is
        510 characters and whose longest domain statement is 596. [measured]
        This raises rather than pre-capping because a CCM statement four times
        its measured maximum means the source or the domain-statement rule
        changed, and silently shortening it would hide that.

        Raises:
            ValueError: If any statement reaches DESCRIPTION_MAX_LENGTH.
        """
        for control in controls:
            if len(control.description) < _DESCRIPTION_LIMIT:
                continue
            raise ValueError(
                f"csa_ccm: {control.control_id} has a statement of "
                f"{len(control.description)} characters, over the "
                f"{_DESCRIPTION_LIMIT}-character limit at which "
                f"BaseParser._sanitize_control truncates the description and "
                f"overwrites full_text with the untruncated text. The pinned "
                f"workbook's longest statement is 596 characters, so this "
                f"means the source or the domain-statement rule moved."
            )

    @classmethod
    def domain_audit_records(
        cls,
        controls: list[Control],
        known_divergences: Mapping[str, Divergence] | None = None,
    ) -> list[dict[str, object]]:
        """What this parser assembled, and the one divergence it left alone.

        A count says a synthesis happened. It does not say what text a link
        now trains on. One record per domain aggregate carries the domain's
        own name as `text_before`, the assembled statement as `text_after`,
        and the member titles that were joined. One record per entry in
        `known_divergences` carries both candidate anchors as text, so a
        reviewer can compare the two rather than take the ruling on trust.

        Raises:
            ValueError: If a divergence names a control this parse did not
                produce, which would leave the link resolving somewhere else
                while the resolution rate still read 1.0000.
        """
        divergences = (
            cls.known_divergences if known_divergences is None
            else known_divergences
        )
        by_id = {control.control_id: control for control in controls}
        records: list[dict[str, object]] = []

        for control in controls:
            if control.hierarchy_level != "domain":
                continue
            titles = list((control.metadata or {}).get("member_titles") or [])
            records.append({
                "kind": "aggregate",
                "control_id": control.control_id,
                "domain_name": control.title,
                "member_count": len(titles),
                "member_titles": titles,
                "statement_chars": len(control.description),
                "text_before": control.title,
                "text_after": control.description,
                "reason": (
                    "The CCM gives a domain no text of its own, so before "
                    "this parser a domain carried its name and nothing else, "
                    "which ProseIndex excludes as a restated title. Its "
                    "statement here is the ordered list of its member control "
                    "titles, assembled by this parser and marked synthetic. "
                    "Concatenating the member specifications instead runs "
                    "1,022 to 4,292 characters, puts 8 of 17 domain anchors "
                    "over MAX_ANCHOR_CHARS, and makes all 17 open with their "
                    "own first member."
                ),
            })

        for section_id, divergence in sorted(divergences.items()):
            target = by_id.get(divergence.resolved_to)
            if target is None:
                raise ValueError(
                    f"csa_ccm: KNOWN_DIVERGENCES names "
                    f"{divergence.resolved_to!r}, which this parse did not "
                    f"produce. The {section_id} link resolves through that "
                    f"control, so a stale entry sends it somewhere else while "
                    f"the resolution rate stays at 1.0000."
                )
            domain = by_id.get(section_id)
            records.append({
                "kind": "wrong_anchor_risk",
                "opencre_section_id": section_id,
                "opencre_section_name": divergence.opencre_section_name,
                "resolved_to": divergence.resolved_to,
                "resolved_by": "section_name",
                "text_before": "" if domain is None else domain.description,
                "text_after": target.description,
                "reason": divergence.reason,
            })
        return records


def main() -> None:
    CsaCcmParser().run()


if __name__ == "__main__":
    main()
