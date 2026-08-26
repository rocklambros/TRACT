"""207 controls and 17 domains, and a domain is its members' subjects.

Measured on the pinned workbook: concatenating member specifications makes 8 of
17 domain anchors exceed MAX_ANCHOR_CHARS and makes all 17 open with their own
first member control. Concatenating member titles makes 0 of either, and runs
163 to 596 characters.

Every string in ROWS is invented, and stays invented. Owner decision D1(b) on
2026-08-26 ruled CSA material redistributable for this project, so the reason
has narrowed rather than disappeared: the corpus may carry CCM specifications,
and a test fixture still has no business quoting a publisher's text to prove a
parser splits columns. A fixture that quotes its source also cannot tell a
parser bug from a source change, because both sides move together. The shape is
the workbook's, the wording is nobody's.

TestSyntheticWorkbook drives parse() and run() against a workbook this file
builds, so the extraction path is covered in CI, where data/raw is absent.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from parsers.parse_csa_ccm import (
    EXPECTED_HEADER,
    KNOWN_DIVERGENCES,
    OPENCRE_TITLE_VARIANTS,
    SHEET_NAME,
    WORKBOOK_NAME,
    CsaCcmParser,
    Divergence,
)
from tract.config import (
    DESCRIPTION_MAX_LENGTH,
    MAX_ANCHOR_CHARS,
    PROCESSED_FRAMEWORKS_DIR,
)
from tract.corpus_report import (
    CURATED_LINKS_PATH,
    SYNTHETIC_TEXT_ORIGIN,
    TEXT_ORIGIN_METADATA_KEY,
)
from tract.parsers.base import BaseParser
from tract.text_selection import prepare_anchor

# The workbook's shape with none of its words. A control row fills all four
# columns; a domain header fills column A alone as "<Full Name> - <CODE>"; the
# title row, the End of Standard trailer and the copyright paragraph are the
# three column-A rows that are not domains.
ROWS: list[tuple[str, str, str, str]] = [
    ('{"specification_name":"Cloud Controls Matrix"}',
     "CLOUD CONTROLS MATRIX v4.1.0", "", ""),
    ("", "", "", ""),
    ("Control Domain", "Control Title", "Control ID", "Control Specification"),
    ("Audit & Assurance - A&A", "", "", ""),
    ("Audit & Assurance", "Ledger Review Cadence And Findings Retention",
     "A&A-01",
     "Publish the cadence on which the ledger is read, and keep each reading "
     "where the next reader can find it without asking anyone."),
    ("Audit & Assurance", "Outside Reader Rotation Between Review Cycles",
     "A&A-02",
     "Rotate the outside reader so that no one reader signs two consecutive "
     "cycles of the same ledger."),
    ("Application & Interface Security - AIS", "", "", ""),
    ("Application & Interface Security", "Build Pipeline Lifecycle", "AIS-04",
     "Run the build through one declared lifecycle, and record which stage "
     "produced each artifact that leaves it."),
    ("Application & Interface Security", "Pipeline Probe Suite", "AIS-05",
     "Keep a probe suite alongside the pipeline and run it on every branch "
     "that can reach a release."),
    ("Application & Interface Security", "Release Gate Signing", "AIS-06",
     "Sign at the release gate, and refuse an artifact whose signature names "
     "no gate this pipeline owns."),
    ("Interoperability & Portability - IPY", "", "", ""),
    ("Interoperability & Portability",
     "Handover Format Register And Escrow Obligations", "IPY-01",
     "Register the formats a handover may use, and say in the agreement who "
     "holds each one in escrow."),
    ("Interoperability & Portability", "Portability Interface Availability",
     "IPY-02",
     "Keep the portability interface reachable for as long as the agreement "
     "says a handover may be requested."),
    ("Infrastructure Security - I&S", "", "", ""),
    ("Infrastructure Security", "Capacity Headroom Planning And Demand Review",
     "I&S-02",
     "Plan headroom against measured demand rather than against the shape of "
     "last year's purchase order."),
    ("Infrastructure Security",
     "Perimeter Traffic Screening Under Change Control", "I&S-09",
     "Screen traffic at the perimeter and keep the screening rules under the "
     "same change control as the hosts behind them."),
    ("End of Standard", "", "", ""),
    ("(c) Copyright 2026 Cloud Security Alliance - All rights reserved.",
     "", "", ""),
]

# The fixture's own shape, so a test that changes ROWS cannot leave a stale
# literal behind in four other tests.
FIXTURE_CONTROLS: int = 9
FIXTURE_DOMAINS: int = 4
FIXTURE_UNITS: int = FIXTURE_CONTROLS + FIXTURE_DOMAINS

# Two of the three v4.0 spellings the real table declares name a control the
# fixture also carries, so the class table is satisfiable here. The third is
# supplied per call where a test needs the table narrowed.
FIXTURE_VARIANTS: dict[str, tuple[str, ...]] = dict(OPENCRE_TITLE_VARIANTS)


def _write_workbook(directory: Path,
                    rows: list[tuple[str, str, str, str]]) -> Path:
    """A workbook shaped like the CCM, worded like nothing in particular."""
    book = openpyxl.Workbook()
    active = book.active
    assert active is not None
    active.title = "Introduction"
    sheet = book.create_sheet(SHEET_NAME)
    for row in rows:
        sheet.append(list(row))
    path = directory / WORKBOOK_NAME
    book.save(path)
    return path


def _parser_for(tmp_path: Path,
                rows: list[tuple[str, str, str, str]],
                name: str = "raw") -> CsaCcmParser:
    """A parser over a fixture workbook, with the real digest pin lifted."""
    raw = tmp_path / name
    raw.mkdir()
    _write_workbook(raw, rows)
    parser = CsaCcmParser(
        raw_dir=raw,
        output_dir=tmp_path / f"out_{name}",
        audit_dir=tmp_path / f"audit_{name}",
    )
    parser.expected_sha256 = None  # type: ignore[misc]
    parser.expected_count = FIXTURE_UNITS  # type: ignore[misc]
    parser.expected_control_rows = FIXTURE_CONTROLS  # type: ignore[misc]
    parser.expected_domains = FIXTURE_DOMAINS  # type: ignore[misc]
    return parser


class TestRowsToControls:
    def test_the_header_row_is_not_a_control(self) -> None:
        controls = CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
        assert "Control ID" not in {c.control_id for c in controls}
        control_rows = [c for c in controls if c.hierarchy_level == "control"]
        assert len(control_rows) == FIXTURE_CONTROLS

    def test_the_three_non_domain_column_a_rows_are_not_domains(self) -> None:
        domains = [
            c for c in CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
            if c.hierarchy_level == "domain"
        ]
        assert sorted(c.control_id for c in domains) == [
            "A&A", "AIS", "I&S", "IPY",
        ]

    def test_a_domain_statement_lists_its_member_titles(self) -> None:
        domain = next(
            c for c in CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
            if c.control_id == "A&A"
        )
        assert domain.description == (
            "Ledger Review Cadence And Findings Retention. "
            "Outside Reader Rotation Between Review Cycles."
        )

    def test_a_domain_statement_does_not_open_its_first_member(self) -> None:
        """The column that would have caught specification concatenation."""
        controls = CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
        domain = next(c for c in controls if c.control_id == "A&A")
        first = next(c for c in controls if c.control_id == "A&A-01")
        assert not domain.description.startswith(first.description[:40])
        assert first.description not in domain.description
        assert len(domain.description) <= MAX_ANCHOR_CHARS

    def test_a_domain_statement_is_marked_synthetic(self) -> None:
        """The trainer must be able to tell a list of subjects from a rule."""
        controls = CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
        domain = next(c for c in controls if c.control_id == "A&A")
        member = next(c for c in controls if c.control_id == "A&A-01")
        assert domain.metadata is not None
        assert (domain.metadata[TEXT_ORIGIN_METADATA_KEY]
                == SYNTHETIC_TEXT_ORIGIN)
        assert TEXT_ORIGIN_METADATA_KEY not in (member.metadata or {})

    def test_a_domain_records_its_member_titles_not_their_ids(self) -> None:
        domain = next(
            c for c in CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
            if c.control_id == "IPY"
        )
        assert domain.metadata is not None
        assert domain.metadata["member_titles"] == [
            "Handover Format Register And Escrow Obligations",
            "Portability Interface Availability",
        ]

    def test_a_control_carries_its_domain_as_parent(self) -> None:
        control = next(
            c for c in CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
            if c.control_id == "I&S-02"
        )
        assert control.parent_id == "I&S"
        assert control.parent_name == "Infrastructure Security"

    def test_a_domain_with_no_members_is_refused(self) -> None:
        """Membership comes from row order, so an empty domain is a reorder."""
        rows = [*ROWS[:4], *ROWS[6:]]
        with pytest.raises(ValueError, match="no controls under it"):
            CsaCcmParser.rows_to_controls(rows, FIXTURE_VARIANTS)

    def test_an_unrecognised_column_a_row_is_refused(self) -> None:
        """A skipped row is a domain that quietly stops existing."""
        rows = [*ROWS[:3], ("Editorial note on the release", "", "", ""),
                *ROWS[3:]]
        with pytest.raises(ValueError, match="is neither a domain header"):
            CsaCcmParser.rows_to_controls(rows, FIXTURE_VARIANTS)

    def test_a_missing_header_row_is_refused(self) -> None:
        rows = [r for r in ROWS if r != EXPECTED_HEADER]
        with pytest.raises(ValueError, match="header rows equal to"):
            CsaCcmParser.rows_to_controls(rows, FIXTURE_VARIANTS)

    def test_a_reordered_header_is_refused(self) -> None:
        """Swapped columns parse 207 controls with title and text exchanged.

        Every count check still passes, so the header is the only place this
        can be caught.
        """
        swapped = ("Control Title", "Control Domain",
                   "Control Specification", "Control ID")
        rows = [swapped if r == EXPECTED_HEADER else r for r in ROWS]
        with pytest.raises(ValueError, match="header rows equal to"):
            CsaCcmParser.rows_to_controls(rows, FIXTURE_VARIANTS)

    def test_a_duplicate_control_id_is_refused(self) -> None:
        rows = [*ROWS, ("Audit & Assurance", "Another Title", "A&A-01",
                        "A second row claiming an identifier that a row "
                        "earlier in the sheet already claimed.")]
        with pytest.raises(ValueError, match="claims control id"):
            CsaCcmParser.rows_to_controls(rows, FIXTURE_VARIANTS)


class TestDeclaredCounts:
    def test_the_declared_total_is_its_two_declared_parts(self) -> None:
        """COUNT_TOLERANCE is 10%, so the band around 224 spans 202 to 246
        and a declared 207 sits inside it. Nothing else here can see a false
        expected_count, because _check_shape is what catches a real loss and
        the band never fires either way.
        """
        assert CsaCcmParser.expected_count == (
            CsaCcmParser.expected_control_rows + CsaCcmParser.expected_domains
        )


class TestAnchorBudget:
    """Ruling R14. BaseParser._sanitize_control rewrites full_text behind the
    parser's back when description exceeds DESCRIPTION_MAX_LENGTH."""

    def test_no_fixture_description_reaches_the_sanitiser_limit(self) -> None:
        controls = CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
        assert max(len(c.description) for c in controls) < DESCRIPTION_MAX_LENGTH

    def test_a_specification_over_the_limit_is_refused(self) -> None:
        rows = [*ROWS, ("Audit & Assurance", "Overlong Statement", "A&A-03",
                        "word " * (DESCRIPTION_MAX_LENGTH // 4))]
        with pytest.raises(ValueError, match="characters, over the "):
            CsaCcmParser.rows_to_controls(rows, FIXTURE_VARIANTS)

    def test_an_overlong_domain_statement_is_refused(self) -> None:
        """The same guard on the side this parser assembles itself."""
        long_title = "Portability " * 40
        rows = [
            *ROWS,
            *[("Interoperability & Portability", f"{long_title}{n}",
               f"IPY-{n:02d}",
               "A statement long enough to be prose and short enough to fit.")
              for n in range(10, 30)],
        ]
        with pytest.raises(ValueError, match="characters, over the "):
            CsaCcmParser.rows_to_controls(rows, FIXTURE_VARIANTS)


class TestTitleVariants:
    def test_a_declared_variant_reaches_the_controls_alt_titles(self) -> None:
        controls = CsaCcmParser.rows_to_controls(
            ROWS, {"AIS-04": ("Secure Application Design and Development",)},
        )
        renamed = next(c for c in controls if c.control_id == "AIS-04")
        untouched = next(c for c in controls if c.control_id == "AIS-05")
        assert renamed.metadata is not None
        assert renamed.metadata["alt_titles"] == [
            "Secure Application Design and Development",
        ]
        assert "alt_titles" not in (untouched.metadata or {})

    def test_a_variant_naming_an_absent_control_is_refused(self) -> None:
        with pytest.raises(ValueError, match="names no control"):
            CsaCcmParser.rows_to_controls(ROWS, {"ZZZ-99": ("Gone upstream",)})

    def test_a_variant_equal_to_the_real_title_is_refused(self) -> None:
        with pytest.raises(ValueError, match="already the control's own title"):
            CsaCcmParser.rows_to_controls(
                ROWS, {"AIS-04": ("build pipeline lifecycle",)},
            )

    def test_an_empty_variant_is_refused(self) -> None:
        with pytest.raises(ValueError, match="declares an empty"):
            CsaCcmParser.rows_to_controls(ROWS, {"AIS-04": ("   ",)})

    def test_every_declared_variant_is_spelled_by_a_curated_link(self) -> None:
        """The tracked half of the ratchet, so it runs without data/raw.

        A declared entry that no curated link spells any more is dead, and a
        dead entry in this table reads as a live one.
        """
        spelled: dict[str, list[str]] = {}
        with CURATED_LINKS_PATH.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                row = json.loads(line)
                if row.get("framework_id") != "csa_ccm":
                    continue
                spelled.setdefault(
                    str(row.get("section_id") or ""), [],
                ).append(str(row.get("section_name") or "").strip())

        assert OPENCRE_TITLE_VARIANTS, "an empty table would pass vacuously"
        for control_id, names in OPENCRE_TITLE_VARIANTS.items():
            assert list(names) == spelled[control_id], control_id

    def test_no_control_is_both_a_variant_and_a_known_divergence(self) -> None:
        """The two tables answer opposite questions about the same link.

        A variant says "this name means that control after all". A divergence
        says "this name means a different control from the one the id names,
        and that is correct". A section_id in both would be self-contradictory.
        """
        assert not set(OPENCRE_TITLE_VARIANTS) & set(KNOWN_DIVERGENCES)


class TestAudit:
    def test_one_record_per_aggregate_plus_the_named_divergence(self) -> None:
        controls = CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
        records = CsaCcmParser.domain_audit_records(controls)
        codes = [r["control_id"] for r in records if r["kind"] == "aggregate"]
        assert codes == ["A&A", "AIS", "IPY", "I&S"]
        divergences = [r for r in records if r["kind"] == "wrong_anchor_risk"]
        assert [r["opencre_section_id"] for r in divergences] == ["IPY"]

    def test_an_aggregate_record_carries_before_and_after_as_text(self) -> None:
        """A count says a synthesis fired. It does not say what a link trains
        on, which is the whole reason this file exists."""
        controls = CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
        record = next(
            r for r in CsaCcmParser.domain_audit_records(controls)
            if r.get("control_id") == "A&A"
        )
        assert record["text_before"] == "Audit & Assurance"
        assert record["text_after"] == (
            "Ledger Review Cadence And Findings Retention. "
            "Outside Reader Rotation Between Review Cycles."
        )
        assert record["member_count"] == 2
        assert record["member_titles"] == [
            "Ledger Review Cadence And Findings Retention",
            "Outside Reader Rotation Between Review Cycles",
        ]

    def test_the_divergence_record_carries_both_candidate_anchors(self) -> None:
        controls = CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
        record = next(
            r for r in CsaCcmParser.domain_audit_records(controls)
            if r["kind"] == "wrong_anchor_risk"
        )
        domain = next(c for c in controls if c.control_id == "IPY")
        control = next(c for c in controls if c.control_id == "IPY-01")
        assert record["resolved_to"] == "IPY-01"
        assert record["resolved_by"] == "section_name"
        assert record["text_before"] == domain.description
        assert record["text_after"] == control.description
        assert record["text_before"] != record["text_after"]

    def test_a_stale_divergence_target_is_refused(self) -> None:
        controls = CsaCcmParser.rows_to_controls(ROWS, FIXTURE_VARIANTS)
        with pytest.raises(ValueError, match="which this parse did not"):
            CsaCcmParser.domain_audit_records(
                controls,
                {"IPY": Divergence("Some name", "IPY-99", "reason")},
            )


class TestSyntheticWorkbook:
    """parse() and run() against a workbook this test writes.

    data/raw is gitignored, so a version of this file that reached parse()
    only through a FileNotFoundError skip would never execute the openpyxl
    path in CI at all.
    """

    def test_parse_reads_the_sheet(self, tmp_path: Path) -> None:
        controls = _parser_for(tmp_path, ROWS).parse()
        assert sorted(c.control_id for c in controls) == [
            "A&A", "A&A-01", "A&A-02", "AIS", "AIS-04", "AIS-05", "AIS-06",
            "I&S", "I&S-02", "I&S-09", "IPY", "IPY-01", "IPY-02",
        ]

    def test_a_missing_ccm_sheet_is_refused(self, tmp_path: Path) -> None:
        raw = tmp_path / "raw_nosheet"
        raw.mkdir()
        book = openpyxl.Workbook()
        active = book.active
        assert active is not None
        active.title = "CAIQ"
        book.save(raw / WORKBOOK_NAME)
        parser = CsaCcmParser(raw_dir=raw, output_dir=tmp_path / "out")
        parser.expected_sha256 = None  # type: ignore[misc]
        with pytest.raises(ValueError, match="has no 'CCM' sheet"):
            parser.parse()

    def test_a_short_sheet_is_refused(self, tmp_path: Path) -> None:
        """The band would accept 202 of 224. The structural check does not."""
        rows = [r for r in ROWS if r[2] != "A&A-02"]
        with pytest.raises(ValueError, match="control rows"):
            _parser_for(tmp_path, rows, name="short").parse()

    def test_a_lost_domain_is_refused(self, tmp_path: Path) -> None:
        """Fifteen of 29 curated links target a bare domain code."""
        rows = [r for r in ROWS
                if not r[0].startswith("Interoperability & Portability - ")]
        parser = _parser_for(tmp_path, rows, name="nodomain")
        parser.expected_control_rows = FIXTURE_CONTROLS  # type: ignore[misc]
        with pytest.raises(ValueError, match="domains, expected"):
            parser.parse()

    def test_a_workbook_that_is_not_the_pinned_one_is_refused(
        self, tmp_path: Path,
    ) -> None:
        parser = _parser_for(tmp_path, ROWS, name="wrongdigest")
        parser.expected_sha256 = "0" * 64  # type: ignore[misc]
        with pytest.raises(ValueError, match="not the pinned"):
            parser.parse()

    def test_run_writes_and_writes_the_audit(self, tmp_path: Path) -> None:
        parser = _parser_for(tmp_path, ROWS, name="run")
        (tmp_path / "out_run").mkdir()
        output = parser.run()
        assert len(output.controls) == FIXTURE_UNITS
        assert [s.path for s in output.source_files] == [WORKBOOK_NAME]
        lines = (tmp_path / "audit_run" / "csa_ccm.jsonl").read_text(
            encoding="utf-8",
        ).splitlines()
        assert len(lines) == FIXTURE_DOMAINS + len(KNOWN_DIVERGENCES)
        assert all(json.loads(line) for line in lines)

    def test_reparse_is_byte_identical(self, tmp_path: Path) -> None:
        parser = _parser_for(tmp_path, ROWS, name="repeat")
        (tmp_path / "out_repeat").mkdir()
        parser.run()
        first = (tmp_path / "out_repeat" / "csa_ccm.json").read_bytes()
        audit = (tmp_path / "audit_repeat" / "csa_ccm.jsonl").read_bytes()
        parser.run()
        assert (tmp_path / "out_repeat" / "csa_ccm.json").read_bytes() == first
        assert (tmp_path / "audit_repeat" / "csa_ccm.jsonl").read_bytes() == audit

    def test_a_title_length_statement_trips_the_prose_floor(
        self, tmp_path: Path,
    ) -> None:
        """Nothing else here proves the floor is set to a non-zero value."""
        rows = [
            (r[0], r[1], r[2], "Short.")
            if r[2] and r[3] and r != EXPECTED_HEADER else r
            for r in ROWS
        ]
        parser = _parser_for(tmp_path, rows, name="titles")
        (tmp_path / "out_titles").mkdir()
        with pytest.raises(ValueError, match="below the declared floor"):
            parser.run()


class TestRealWorkbook:
    """The pinned source. Skipped where data/raw is absent, which is CI."""

    @pytest.fixture(scope="class")
    def controls(self) -> list[Any]:
        parser = CsaCcmParser()
        try:
            return parser.parse()
        except FileNotFoundError:
            pytest.skip("data/raw is gitignored and absent in this checkout")

    def test_run_writes_from_the_real_workbook(self, tmp_path: Path) -> None:
        parser = CsaCcmParser(
            output_dir=tmp_path, audit_dir=tmp_path / "audit",
        )
        try:
            output = parser.run()
        except FileNotFoundError:
            pytest.skip("data/raw is gitignored and absent in this checkout")
        assert len(output.controls) == 224
        domains = [c for c in output.controls if c.hierarchy_level == "domain"]
        assert len(domains) == 17
        assert [s.path for s in output.source_files] == [WORKBOOK_NAME]

    def test_the_prose_fraction_clears_the_declared_floor(
        self, controls: list[Any],
    ) -> None:
        """222 of 224 clear HONEST_PROSE_MIN_CHARS. IAM-07 is 58 characters
        and STA-06 is 43, so a floor of 1.0 would refuse correct output."""
        fraction = BaseParser.honest_prose_fraction(controls)
        assert fraction == pytest.approx(222 / 224)
        assert fraction >= CsaCcmParser.min_prose_fraction
        assert CsaCcmParser.min_prose_fraction > 221 / 224

    def test_no_shipped_description_reaches_the_sanitiser_limit(
        self, controls: list[Any],
    ) -> None:
        """Ruling R14, asserted on the source rather than assumed."""
        assert max(len(c.description) for c in controls) < DESCRIPTION_MAX_LENGTH

    def test_every_domain_statement_fits_the_encoder_budget(
        self, controls: list[Any],
    ) -> None:
        lengths = [
            len(c.description) for c in controls
            if c.hierarchy_level == "domain"
        ]
        assert len(lengths) == 17
        assert max(lengths) <= MAX_ANCHOR_CHARS
        # Specification concatenation runs 1,022 to 4,292 and puts 8 of the 17
        # over budget. Title concatenation runs 163 to 596. [measured]
        assert (min(lengths), max(lengths)) == (163, 596)

    def test_the_anchors_share_no_leading_prefix(
        self, controls: list[Any],
    ) -> None:
        """Ruling R13, measured rather than assumed.

        The Top 10's ten anchors opened with 364 byte-identical characters of
        a statistics table. These share nothing, so there is no boilerplate to
        strip and no removal to audit.
        """
        anchors = sorted(prepare_anchor(c.description)[0] for c in controls)
        shortest, longest = anchors[0], anchors[-1]
        shared = 0
        while (shared < len(shortest) and shared < len(longest)
               and shortest[shared] == longest[shared]):
            shared += 1
        assert shared == 0

    def test_the_variant_table_is_exactly_what_the_source_renamed(
        self, controls: list[Any],
    ) -> None:
        """The other half of the ratchet, and the half that needs the source.

        A curated name that stops matching the title of the control its own id
        names, and is not declared here or ruled a divergence, fails. A
        declared entry the source no longer renames fails too.
        """
        titles = {c.control_id: c.title for c in controls}
        observed: dict[str, list[str]] = {}
        with CURATED_LINKS_PATH.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                row = json.loads(line)
                if row.get("framework_id") != "csa_ccm":
                    continue
                section_id = str(row.get("section_id") or "")
                name = str(row.get("section_name") or "").strip()
                title = titles.get(section_id)
                if title is None or section_id in KNOWN_DIVERGENCES:
                    continue
                if name.lower() != title.lower():
                    observed.setdefault(section_id, []).append(name)

        assert observed == {
            key: list(names) for key, names in OPENCRE_TITLE_VARIANTS.items()
        }

    def test_the_written_artifact_is_tracked(self) -> None:
        """csa_ccm is tracked since 2026-08-26, owner decision D1(b).

        Inverted rather than deleted. This assertion is the pair to the tier
        membership in tract.config: OVERLAY_FRAMEWORK_IDS decides what the
        merge withholds from the tracked corpus, and .gitignore decides what a
        fresh clone receives, and the two disagreeing is a real failure mode
        rather than a hypothetical. They did disagree for one commit, when the
        tier changed and the ignore line did not, and a fresh clone got a
        framework the corpus described and the tree could not produce.

        So the direction flips with the ruling and the check stays. If CSA's
        terms are ever re-read the other way, this fails and says so.
        """
        import subprocess

        path = PROCESSED_FRAMEWORKS_DIR / "csa_ccm.json"
        result = subprocess.run(
            ["git", "check-ignore", "-q", str(path)],
            capture_output=True,
            cwd=Path(__file__).resolve().parent.parent,
        )
        assert result.returncode != 0, (
            f"{path} is gitignored, but csa_ccm left OVERLAY_FRAMEWORK_IDS on "
            f"owner decision D1(b). The tier and .gitignore have to agree, or "
            f"a fresh clone gets a framework the corpus describes and the tree "
            f"cannot produce."
        )
